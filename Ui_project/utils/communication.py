from openpyxl import Workbook, load_workbook
import serial
from pathlib import Path
import time

MAX_NO_OF_INPUTS = 100
MAX_NO_OF_OUTPUTS = 10

filename = 'CommTestData3.xlsx'
loc = Path(filename)
wb = load_workbook(loc)

sheet = wb.active
print(sheet.cell(row=1, column=1).value)

NO_OF_ROWS = sheet.max_row-1
##NO_OF_ROWS = 50


for i in range(1, MAX_NO_OF_INPUTS+2):
    if sheet.cell(row = 1, column = i).value == None:
        numberOfInputs = i-1
        print("The number of inputs is ", numberOfInputs)
        break

for i in range(numberOfInputs+2, MAX_NO_OF_INPUTS+MAX_NO_OF_OUTPUTS+3):
    if sheet.cell(row = 1, column = i).value == None:
        numberOfOutputs = i-numberOfInputs-2
        print("The number of outputs is ", numberOfOutputs)
        break
    
arduino = serial.Serial('COM9', 115200 , timeout=.1)
for t in range(2, NO_OF_ROWS+2):
    if (t ==2):
        startingTime = int(round(time.time() * 1000))
    elif (t == NO_OF_ROWS+1):
        endingTime = int(round(time.time() * 1000))
    print(t)
    for i in range(1,numberOfInputs+1):
        toPrint = str(sheet.cell(row = t, column = i).value)
##        print(toPrint)
        toSend = bytearray(toPrint+chr(0)+chr(3), encoding="utf-8")
        arduino.write(toSend)
        
    arduino.write(bytearray(chr(4), encoding="utf-8"))
##    String outputs[numberOfOutputs]
    for i in range(numberOfInputs+2,numberOfInputs+numberOfOutputs+2):
        time.sleep(.005)
        outputRead = ""
        while (arduino.inWaiting()==0):
                pass
        while (arduino.inWaiting()>0):
            receivedByte = arduino.read(1)
##            print(receivedByte)
            if(receivedByte == bytearray.fromhex('03')):
                break;
            outputRead = outputRead + str(receivedByte, 'utf-8')
##            print(outputRead[-1:].encode("utf-8").hex())
            
##        print("I read the string "+outputRead)
        cellref = sheet.cell(row = t, column = i)
        cellref.value = float(outputRead)
    while (arduino.inWaiting()==0):
        pass
    while (arduino.inWaiting()>0):
        receivedByte = arduino.read(1)
        if(receivedByte!= bytearray.fromhex('04')):
            print("An error has accured with serial communication, please run the program again")
            quit()
wb.save(filename[:-5]+"Updated.xlsx")
print("DONE! Took me ", (endingTime-startingTime)/NO_OF_ROWS, " milliseconds per row on average")

